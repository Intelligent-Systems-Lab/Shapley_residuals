import re
import pandas as pd
import numpy as np
from trans import instance_data_to_subset_list
from feature_name import feature_name
from partial_gradient_square import partial_gradient_square
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color

def save_single_data(index_list):
    '''subset_length=1的scaled norm存成csv檔'''
    _data_allidx = []
    for idx in index_list:
        _data = []
        data = instance_data_to_subset_list(idx)
        ###單一特徵
        for _,value in data[0].items():
            _data.append(value)
        _data_allidx.append(_data)
    df = pd.DataFrame(data=_data_allidx,index=index_list,columns=feature_name)
    df.to_csv(f'./csv_data_saved/instance{index_list}_單一特徵scaled_norm.csv')
    print('csv文件已保存!')
    return df

def save_len2_data(index_list):
    #tt_num = len(index_list)
    _data_allidx = []
    for idx in index_list:
        data = instance_data_to_subset_list(idx)
        feature_num = len(data[0])
        _data = np.zeros((feature_num,feature_num))
        for key, value in data[1].items():
            cord = [int(num) for num in re.findall(r'\d+', key)] 
            _data[cord[0]][cord[1]] = value
            _data[cord[1]][cord[0]] = value
        values = list(data[0].values())
        np.fill_diagonal(_data,values)
        df = pd.DataFrame(data=_data,index=feature_name,columns=feature_name)
        df.to_csv(f'./csv_data_saved/instance{idx}_兩兩特徵scaled_norm.csv')
        _data_allidx.append(df)
    
    print(f'已保存{len(index_list)}個文件!')
    return _data_allidx
##todo
def save_partial_gradient_square(index_list):
    a = [partial_gradient_square[idx] for idx in index_list]
    df = pd.DataFrame(data=a,index=index_list,columns=feature_name)
    df.to_csv(f'./csv_data_saved/instance{index_list}_partial_gradient_square.csv')
    print('已保存partial gradient square')
    return df

def save_len2_data_with_color(index_list):
    wb = Workbook()
    ws = wb.active
    df_list  = save_len2_data(index_list)
    ##寫單一Dataframe
    def write_df_with_empty_lines(df, row_start, col_start, empty_rows, idx):
        cell = ws.cell(row=row_start-1,column=col_start-1,value=f'index_{idx}')
        for c_idx, col in enumerate(df.columns, start=col_start):  # 列索引從列2
            cell = ws.cell(row=row_start-1, column=c_idx, value=col)
        for r_idx, idx in enumerate(df.index, start=row_start):  # 行索引從行2
            cell = ws.cell(row=r_idx, column=col_start-1, value=idx)
        for i, col in enumerate(df.columns, start=col_start):  
            for j, index in enumerate(df.index, start=row_start):  
                if(i-col_start)<(j-row_start):
                    break
                if (i-col_start)==(j-row_start):
                    cell = ws.cell(row=j, column=i)
                    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                if df.iloc[j-row_start, i-col_start] < df.iloc[j-row_start, j-row_start] and df.iloc[j-row_start, i-col_start] < df.iloc[i-col_start, i-col_start]:
                    # 著色
                    cell = ws.cell(row=j, column=i)
                    cell.fill = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
        for r_idx, row in enumerate(df.iterrows(), start=row_start):
            for c_idx, value in enumerate(row[1], start=col_start):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)  
        for _ in range(empty_rows):
            ws.append([])  # 添加空行
    r_s = 2
    c_s = 2
    e_r = 3
    for i,idx in enumerate(index_list):
       df = df_list[i]
       write_df_with_empty_lines(df,row_start=r_s,col_start=c_s,empty_rows=e_r,idx=idx)
       r_s += df.shape[0] + e_r + 1

    # 保存 Excel 文件
    print('文件存檔成功!')
    wb.save(f"instance{index_list}_兩兩特徵.xlsx")
    return 0

'''for df in df_list:
        ##寫index
        for c_idx, col in enumerate(df.columns, start=2):  # 列索引從列2
            cell = ws.cell(row=1, column=c_idx, value=col)

        for r_idx, idx in enumerate(df.index, start=2):  # 行索引從行2
            cell = ws.cell(row=r_idx, column=1, value=idx)
        #寫數值
        for i, col in enumerate(df.columns, start=2):  
            for j, index in enumerate(df.index, start=2):  
                if i==j:
                    cell = ws.cell(row=j, column=i)
                    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                if df.iloc[j-2, i-2] < df.iloc[j-2, j-2] and df.iloc[j-2, i-2] < df.iloc[i-2, i-2]:
                    # 著色
                    cell = ws.cell(row=j, column=i)
                    cell.fill = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
        for r_idx, row in enumerate(df.iterrows(), start=1):
            for c_idx, value in enumerate(row[1], start=1):
                cell = ws.cell(row=r_idx+1, column=c_idx+1, value=value)  '''
    


